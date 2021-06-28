from tkinter import *
import threading
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import random
import os
import webbrowser  
import tkinter as tk

#-----------------------------------ToolTip class--------------------------------------------------#
class CreateToolTip(object):
    """
    create a tooltip for a given widget
    """
    def __init__(self, widget, text='widget info'):
        self.waittime = 200     #miliseconds
        self.wraplength = 180   #pixels
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        self.widget.bind("<ButtonPress>", self.leave)
        self.id = None
        self.tw = None

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(self.waittime, self.showtip)

    def unschedule(self):
        id = self.id
        self.id = None
        if id:
            self.widget.after_cancel(id)

    def showtip(self, event=None):
        x = y = 0
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        # creates a toplevel window
        self.tw = tk.Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(self.tw, text=self.text, justify='left',
                       background="#ffffff", relief='solid', borderwidth=1,
                       wraplength = self.wraplength)
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tw
        self.tw= None
        if tw:
            tw.destroy()

#-----------------------------------Necessary def--------------------------------------------------#
        
def tem_list(A, B):
    tem_list = list()
    for i in A:
        if(i!=B):
            tem_list.append(i)
    return tem_list

def random_pick(collection,slot):
    result = list()
    tempory_list = collection
    count = 0
    while count<slot:
        result.append(random.choice(tempory_list))
        tempory_list = tem_list(tempory_list, result[count])
        count+=1
    return result

def mazii_search():
    word = canvas.itemcget(card_title,'text')
    url = 'https://mazii.net/search/word?dict=javi&query='+word+'&hl=vi-VN'
    webbrowser.open(url, new=0, autoraise=True)

def google_image_search():
    word = canvas.itemcget(card_title,'text')
    url = 'https://www.google.com.vn/search?q='+word+'&hl=vi&source=lnms&tbm=isch'
    webbrowser.open(url, new=0, autoraise=True)

#-----------------------------------def with flash card handle--------------------------------------------------#

def mix_card():
    global random_list, current_index
    random_list = random_pick(list_index,number_words)
    current_index=1

def flip_card():
    global isBack
    if isBack:
        canvas.itemconfig(card_background, image=card_front_img)
        canvas.itemconfig(card_title, fill="black")
        canvas.itemconfig(card_title, text = ws.cell(random_list[current_index-1],1).value)
        
        isBack = False
    else:        
        canvas.itemconfig(card_background, image=card_back_img)
        canvas.itemconfig(card_title, fill="white", text = ws.cell(random_list[current_index-1],2).value)
        #Show mean of word
        canvas.itemconfig(card_word, fill="white",text = ws.cell(random_list[current_index-1],3).value)
        if ws.cell(random_list[current_index-1],4).value == None:
            canvas.itemconfig(card_kanji, text ="")
        else:
            canvas.itemconfig(card_kanji, fill="white",text = ws.cell(random_list[current_index-1],4).value)
        isBack = True
        
def next_word():
    global current_index, isBack
    if current_index +1 > number_words:
        mix_card()
        current_index=0
    if isBack:
        canvas.itemconfig(card_background, image=card_front_img)
        canvas.itemconfig(card_title, fill="black")
        isBack = False
    current_index+=1
    canvas.itemconfig(card_title, text = ws.cell(random_list[current_index-1],1).value)
    card_No_textvariable.set(str(current_index) + "/" + str(number_words))    
   
def back_word():
    global current_index, isBack
    if current_index <2:
        pass
    else:
        if isBack:
            canvas.itemconfig(card_background, image=card_front_img)
            canvas.itemconfig(card_title, fill="black")
        current_index-=1    
        canvas.itemconfig(card_title, text = ws.cell(random_list[current_index-1],1).value)
        card_No_textvariable.set(str(current_index) + "/" + str(number_words))

    isBack = False

#-------------------------Modify soucre of word to learn---------------------------------------#

def select_workbook():
    global link,ws,number_words,list_index,wb
    link = filedialog.askopenfilename(initialdir = "/",
                                        title = "Select A File",
                                        filetype = (('xlsx', '*.xlsx'),
                                                    ('All', '*.*')))
    try:
        wb = load_workbook(link)
        ws = wb.active 
        number_words = min(len(ws['A']),len(ws['B']))
        list_index = list()
        for i in range(1,number_words+1):
            list_index.append(i)
        mix_card()
        next_word()
    except InvalidFileException:
        pass
    
def select_worksheet():
    global list_of_sheet,v,sheet_select_window

    def sheet_select_window_closing():       
        global ws,number_words,list_index,current_index
        ws = wb[list_of_sheet[v.get()]]
        number_words = min(len(ws['A']),len(ws['B']))
        list_index = list()
        for i in range(1,number_words+1):
            list_index.append(i)
        mix_card()
        current_index = 0
        next_word()

        sheet_select_window.destroy()
    
    list_of_sheet = wb.sheetnames
    sheet_select_window = Toplevel(window)
    sheet_select_window.resizable(False, False)
    sheet_select_window.title("Select sheet")
    sheet_select_window.config(padx =50, pady = 50)
    sheet_select_window.protocol("WM_DELETE_WINDOW", sheet_select_window_closing)
    v = IntVar()
    count = 0
    for sheet in  list_of_sheet:
        Radiobutton(sheet_select_window, 
                    text=sheet,
                    padx = 20, 
                    variable=v, 
                    value=count).pack()
        count+=1

#-----------------------EXERCISES-----------------------------------#
        
def create_exercise():
    global random_question_list,user_answer_list
    random_question_list = list()
    max_range = number_words+1
    user_answer_list=list()
    for i in range(1,max_range):
        random_tem_list = random_pick(tem_list(list_index,i),3)
        random_tem_list.append(i)
        random_question_list.append(random_pick(random_tem_list,4))
        random_question_list[i-1].append(i)
        user_answer_list.append(-1)
  
##        print(random_question_list[i-1])

def next_question():
    global current_question,v

    if current_question == number_words:
        current_question =1
    else:
        current_question+=1

    if user_answer_list[current_question-1]==-1:
        v.set(None)
    else:
        v.set(user_answer_list[current_question-1])
    
    for i in range(4):
        list_of_radiobutton[i].config(text=ws.cell(random_question_list[current_question-1][i],3).value)                                     
    question_textvariable.set(ws.cell(random_question_list[current_question-1][4],1).value)
    question_No_textvariable.set(str(current_question) + "/" + str(number_words))

   
def back_question():
    global current_question,v
    
    if current_question ==1:
        pass
    else:
        current_question-=1

    if user_answer_list[current_question-1]==-1:
        v.set(None)
    else:
        v.set(user_answer_list[current_question-1])
        
    for i in range(4):
        list_of_radiobutton[i].config(text=ws.cell(random_question_list[current_question-1][i],3).value)
                                      
    question_textvariable.set(ws.cell(random_question_list[current_question-1][4],1).value)
    question_No_textvariable.set(str(current_question) + "/" + str(number_words))
    

##def play(file):
##    if file == 'correct':
##        playsound("sounds/correct.mp3")
##    else:
##        playsound("sounds/wrong.mp3")
    

def select_answer(value):    
    user_answer_list[current_question-1]=value
    if random_question_list[current_question-1][value-1] == random_question_list[current_question-1][-1]:
        list_of_radiobutton[value-1].config(selectcolor = 'green')
##        print("Selected answer is right")
    else:
##        print("Selected answer is wrong")
        list_of_radiobutton[value-1].config(selectcolor = 'red')
        globals()['current_score']-=1
        globals()['score_variable'].set( 'Your score now is: ' + str(globals()['current_score']))





    
def exercise():
    global current_question,question_No_textvariable,question,random_question_list,list_of_radiobutton,question_textvariable,current_score,score_variable,v
    newWindow = Toplevel(window)
    newWindow.geometry("800x526")
    newWindow.resizable(False, False)
    newWindow.title("Exercise")
    newWindow.config(padx =50, pady = 50, bg = BACKGROUND_COLOR)
    newWindow.bind('<Key>',exercise_key)

    newWindow.grid_columnconfigure(0, weight=1)
    newWindow.grid_columnconfigure(1, weight=3)
    newWindow.grid_columnconfigure(2, weight=1)

    create_exercise()
    current_question = 1    
    question_textvariable = StringVar()
    question = Label(newWindow,
                     font=("Ariel", 55),
                     highlightthickness=0,
                     bg = BACKGROUND_COLOR,
                     bd = 0,
                     textvariable =question_textvariable).grid(row=0, column=1, pady =10)

    question_textvariable.set(ws.cell(random_question_list[0][4],1).value)

    v = IntVar()
    list_of_radiobutton = [1,2,3,4]
    for i in range(4):
##        if random_question_list[0][i] == random_question_list[0][-1]:
##            list_of_radiobutton[i] = Radiobutton(newWindow,
##                                     text = ws.cell(random_question_list[0][i],3).value,
##                                     width=30,                                             
##                                     font=("Times", 20), variable = v,
##                                     highlightthickness=0, bd = 0,
##                                     value = i+1,
##                                     indicator = 0,
##                                     selectcolor = 'green',
##                                     command=lambda:select_answer(v.get()))
##        else:

        list_of_radiobutton[i] = Radiobutton(newWindow,
                                             text = ws.cell(random_question_list[0][i],3).value,
                                             width=30,                                             
                                             font=("Times", 20), variable = v,
                                             highlightthickness=0, bd = 0,
                                             value = i+1,
                                             indicator = 0,
                                             selectcolor = 'red',
                                             command=lambda:select_answer(v.get()))
        

        list_of_radiobutton[i].grid(row=i+1, column=1,pady =10)

    next_question_bt = Button(newWindow,image=next_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = next_question)
    next_question_bt.grid(row=1, column=2, rowspan = 4)
    
    back_question_bt = Button(newWindow,image=back_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = back_question)
    back_question_bt.grid(row=1, column=0, rowspan = 4)   
        
    question_No_textvariable = StringVar()
    question_No = Label(newWindow,font=("Ariel", 18),bg = BACKGROUND_COLOR, textvariable = question_No_textvariable, width=7)
    question_No_textvariable.set(str(current_question) + "/" + str(number_words))
    question_No.grid(row=6, column=1, padx = 20, ipady = 10)

    score_variable = StringVar()
    current_score = number_words
    score = Label(newWindow,font=("Ariel", 28),bg = BACKGROUND_COLOR, textvariable = score_variable, width=18)
    score_variable.set( 'Your score now is: ' + str(current_score))
    score.grid(row=7, column=1, padx = 20, ipady = 10)
            
        
        
##def main_closing():
##    if tk.messagebox.askokcancel("Quit", "Do you want to quit?"):
##        window.destroy()

def key(event):
    if event.keysym == 'Right':
        next_word()
    if event.keysym == 'Left':
        back_word()
def exercise_key(event):
    if event.keysym == 'Right':
        next_question()
    if event.keysym == 'Left':
        back_question()
      
#Initial run
link = "data/japan_words.xlsx"
#Gather information from wookbook
wb = load_workbook(link)
ws = wb.active    
number_words = min(len(ws['A']),len(ws['B']))
list_index = list()
for i in range(1,number_words+1):
    list_index.append(i)
current_index=1
mix_card()
#Card is in front face
isBack = False


BACKGROUND_COLOR = "#B1DDC6"
BUTTON_BACKGROUND_COLOR = "white"
window = Tk()
window.resizable(False, False)
window.title("Flashy")
window.config(padx =50, pady = 50, bg = BACKGROUND_COLOR)
##window.protocol("WM_DELETE_WINDOW", main_closing)
window.bind('<Key>',key)

menu_bar = Menu(window)
window.config(menu=menu_bar)

select_menu = Menu(menu_bar,tearoff=False)
menu_bar.add_cascade(label="Select", menu=select_menu)
select_menu.add_command(label="Workbook",command=select_workbook)
select_menu.add_command(label="Sheet",command=select_worksheet)
select_menu.add_command(label="Exercise",command=exercise)
select_menu.add_separator()
select_menu.add_command(label="Exit", command=window.destroy)

canvas = Canvas(width = 800, height = 526)
card_front_img = PhotoImage(file="images/card_front.png")
card_back_img = PhotoImage(file="images/card_back.png")

card_background = canvas.create_image(400, 263, image = card_front_img)
canvas.grid(row=1, column=2, columnspan = 3)
canvas.config(bg = BACKGROUND_COLOR, highlightthickness=0)

card_kanji = canvas.create_text(400,150, font=("Ariel", 30))
card_title = canvas.create_text(400,236, font=("Ariel", 55), text = ws.cell(random_list[current_index-1],1).value)
card_word = canvas.create_text(400,350, font=("Ariel", 30, "bold"), width=700, justify='center')

mazii_search_image= PhotoImage(file="images/Mazii_search.png")
mazii_search_bt = Button(image=mazii_search_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = mazii_search)
mazii_search_bt.grid(row=2, column=5)
CreateToolTip(mazii_search_bt,'Search word in Mazii')

google_search_image= PhotoImage(file="images/Google_image.png")
google_search_bt = Button(image=google_search_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = google_image_search)
google_search_bt.grid(row=2, column=0, padx=10)
CreateToolTip(google_search_bt,'Search for word in Google Image')

next_image= PhotoImage(file="images/Next.png")
next_bt = Button(image=next_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = next_word)
next_bt.grid(row=1, column=5)

back_image= PhotoImage(file="images/Back.png")
back_bt = Button(image=back_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = back_word)
back_bt.grid(row=1, column=0, padx=40)

cross_image= PhotoImage(file="images/wrong.png")
unknown_bt = Button(image=cross_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = flip_card)
unknown_bt.grid(row=2, column=2,sticky="E")
CreateToolTip(unknown_bt,'Unknown')

check_image= PhotoImage(file="images/right.png")
known_bt = Button(image=check_image, highlightthickness=0, bg = BACKGROUND_COLOR, bd = 0, command = next_word)
known_bt.grid(row=2, column=4,sticky="W")  
CreateToolTip(known_bt,'Known')

card_No_textvariable = StringVar()
card_No = Label(window,font=("Ariel", 18),bg = BACKGROUND_COLOR, textvariable = card_No_textvariable, width=7)
card_No_textvariable.set(str(current_index) + "/" + str(number_words))
card_No.grid(row=2, column=3, padx = 20, ipady = 10, sticky="N")
window.mainloop()
